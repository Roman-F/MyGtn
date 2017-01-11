# coding=utf-8
from django.shortcuts import render, get_list_or_404
from django.http import HttpResponseRedirect, HttpResponse
from django.db.models import get_model

from openpyxl.styles import Font
from django.conf import settings
import openpyxl
import openpyxl.utils
import random
import time
import os
import models
import json
from bson import json_util
import datetime
from django.core.servers.basehttp import FileWrapper

# Create your views here.
PATH_NATURAL_PERSON = 'AppGtn/register/templates/'

# Словарь соответствия имени модели шаблону отображения реестра
DICT_TEMPLATES_FOR_REGISTER_DISPLAY = {'EntityNaturalPerson': 'register_natural_person.html',
                                       'EntityVehicle': 'register_vehicles.html',
                                       'VehicleColor': 'vehicle_color.html',
                                       'VehicleBrand': 'vehicle_brand.html/',
                                       'Country': 'country.html'
                                       }

# Словарь соответствия имени модели и адресу реестра, ее отображающего
#TODO: придумать как использовать \register\urls.py , там эта информация уже есть (нарушен принцип DRY)
DICT_URL_PATH_TO_REGISTER = {'EntityNaturalPerson': '/register/fl/',
                             'EntityVehicle': '/register/tech/',
                             'VehicleColor': '/register/vehiclecolor/',
                             'VehicleBrand': '/register/vehiclebrand/',
                             'Country': '/register/country/'
                             }


def get_response_to_unload_file (path_to_file, name_file_for_user = '', remove_file = False):

    '''
        Функция читает полученный файл в двоичном режиме.
        По прочитанным данным формирует и возвращает объект HttpResponse.
        Удаляет прочитанный файл, если remove_file = True

        :param path_to_file: Путь к файлу, по которому нужно сформровать HttpResponse.
        :param name_file_for_user: Имя файла, которое будет отображаться пользователю при скачивании
        :param remove_file: если True - то прочитаный файл будет удален
        :return: объект HttpResponse
    '''

    if name_file_for_user == '':
        name_file_for_user = os.path.basename(path_to_file)

    # формируем response с файлом для отдачи пользователю
    with open(path_to_file, 'rb') as myfile:
        response = HttpResponse(myfile.read(), content_type='application')
        response['Content-Disposition'] = 'attachment; filename= ' + name_file_for_user

    if remove_file:
        os.remove(myfile.name)

    return response


def appgtn_register(request,model):
    """
        Общее представление для реестров и справочников
    """
    #TODO: обработать случай когда модель не получена (model is None)
    #TODO: обработать случай когда в БД нет данных

    NAMES_NO_DISPLAYED_FIELDS = ['id','created_date','modifided_date','deleted_date']

    list_fields = list(model._meta.get_fields_with_model())

    map(lambda x: list_fields.remove(x) if x[0].name in NAMES_NO_DISPLAYED_FIELDS else None, list_fields[:])

    list_dict_values = model.objects.all().values(*(map(lambda x: x[0].name, list_fields)))

    table_headers = tuple(x[0].verbose_name for x in list_fields)

    # Формируем кортеж кортежей со значениями из БД, в порядке следования полей в list_fields
    tuple_tuples_values = tuple(tuple(dict_values[x[0].name] for x in list_fields) for dict_values in list_dict_values)

    path_to_template = "base_register.html"

    return render(request, path_to_template, {"table_headers": table_headers,
                                            "tuple_tuples_values": tuple_tuples_values,
                                            "model": model.__name__})

def appgtn_form_register(request):
    """
    Добавление новой и изменение имеющейся записи реестра
    """
    #TODO: для изменения записей необходимо определить как пользователь будет выбирать нужную запись

    name_model = request.POST.get('NameRegisterModel', 'Модель не найдена')
    model_used = get_model('AppGtn', name_model)
    id_record = request.POST.get('id_record', '')

    if request.POST.get('save', ''):

        if id_record:
            data_model = model_used.objects.get(pk=int(id_record))
            form = model_used.get_class_model_form()(request.POST, instance=data_model)
        else:
            form = model_used.get_class_model_form()(request.POST)

        if form.is_valid():
            form.save()
            return HttpResponseRedirect(DICT_URL_PATH_TO_REGISTER[name_model])

    elif id_record:

        data_model = model_used.objects.get(pk=int(id_record))
        meta = model_used._meta
        dict_to_json = {fields.name: data_model.__getattribute__(fields.name) for fields in meta.fields}
        form = model_used.get_class_model_form()(initial=dict_to_json)

    else:
         form=model_used.get_class_model_form()()

    # path_to_template = DICT_TEMPLATES_FOR_FORMS_CHANGES[name_model]
    path_to_template = 'base_form_register.html'

    return render(request, path_to_template, {'form': form
                                              ,'id_record': id_record
                                              ,'path_back':DICT_URL_PATH_TO_REGISTER[name_model]
                                              ,'model': request.POST.get('NameRegisterModel',
                                                                        'Модель не передана')})


# def app_change_register_entity (request):
#     """
#     Задача решаемая данной вьюхой:
#     При нажатии на кноку "Изменить" в любом реесре, происходит определение, что за реестр (какая модель используется),
#     далее открывается форма изменение выбранной записи,
#     в которой вносятся измения и после нажатия на кнопку "Сохранить" записываются в БД,
#     форма изменения при этом закрывается и происходит обновление реестра
#
#     Требования:
#     1) одна вьюха на все реестры
#     2) единообразный способ получения данных о записи и ее передачи в форму
#
#     Вопросы которые необходимо решить
#     1) как определять какая запись выбрана
#
#     Примерный алгоритм работы
#
#     1) Получаем из реестра запрос с данными модели и идентификатором выбранной записи
#     2) Определяем модель реестра
#     3) Получаем QuerySet выбранной записи
#     4) формируем из данных JSON-пакет
#     5) зная имя модели определяем в коллекции нужный шаблон
#     6) формируем RESPONSE с найденным шаблоном и собранным JSON и отдаем его пользователю
#
#     """

def appgtn_in_file_register_natural_person(request):
    # Задача решаемая данной вьюхой:
    # При нажатии на кнопку "Выгрузить" в любом реестре, происходит определение что за реестр (какая модель используеся)
    # и происходит выгрузка всех данных данного реестра.
    #
    # Требования:
    # 1) Одна вьюха для обработки нажатия "Выгрузить" в любом реестре
    # 2) Выгрузка происходит в создаваемый при выгрзке файл формата xlsx
    # 3) Файл создается с нуля на основе данных выгружаемой модели и ее структуры
    # 4) Файлы должны быть отформатированы по единому стандарту (читабельность, ширина, выделнные заголовки)

    # Вопросы:
    # как организовать универсальную передачу параметров
    #Для это в шаблон реестра необходимо передавать имя модели и далее это имя возвращать во вьюху выгрузки файла
    # как реализовать получение данных из нужной модели
    #Во вьюхе выгрузки файла с помощью get_model обращаться к модели и получить объект модели
    #Далее работая с моделью получать нужные данные
    # как реализовать создание и запись в файл
    # для этого необходимо исползовать стороннюю бибилиотеку по работе с файлами XLS
    # используем библиотеку openpexl
    # Примерный алгоритм
    #     Создаем новую книгу
    #     Получаем из модели человеко-понятные названия столцов (verbose_name)
    #     Записываем человекопонятные названия в файл в первую строку (это будут заголовки)
    #     Записываем в файл по колоночно данные из БД (1 строка - один объект БД)
    #     Оформляем файл: первые столбцы жирным, устанавливаем ширину колонок
    #     Отдаем файл пользователю
    # как реализовать отдачу файла пользователю
    # для этого открываем сохраненную книгу как файл
    # затем зашиваем данные файла в объект тип RESPONSE, донастраиваем и отдаем пользователю

    name_model = request.POST.get('NameRegisterModel', 'Модель не найдена')
    model_for_file = get_model('AppGtn', name_model)
    data_model = model_for_file.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active

    #получаем список имен полей модели, для использования далее как заголовков столбцов
    #а также определяем по заголовкам первичную ширину колонок
    #и делаем заголовок жирным
    i = 1
    column_widths = []
    data_field_names = []
    for data in model_for_file._meta.fields:
        ws.cell(row=1, column=i, value=data.verbose_name).font = Font(bold=True)
        data_field_names.append(data.name)
        column_widths.append(len(data.verbose_name) + 1)
        i = i + 1

    #получаем значение полей из БД, и пишем их в эксель
    #а также определяем максимальную ширину колонок
    i = 1
    j = 2
    for exemplar in data_model:
        for name_field in data_field_names:
            field_value = exemplar.__getattribute__(name_field)
            ws.cell(row=j, column=i, value=field_value)
            column_widths[i - 1] = max(column_widths[i - 1], len(str(field_value))) + 1
            i = i + 1
        j = j + 1

    #проставляем ширину колонок в эксель

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column_width


    #Проводим сохранение и выгрузку пользователю файла с данными
    name_file = name_model + str(random.randint(1000, 9999)) + '.xlsx'
    full_name_file = settings.STATIC_ROOT + name_file

    wb.save(full_name_file)

    with open(full_name_file, 'rb') as myfile:
        response = HttpResponse(myfile.read(), content_type='application')
        response['Content-Disposition'] = 'attachment; filename=' + name_file
    myfile.close()

    return response

def appgtn_upload_file_import (request):

    '''
        Вьюха выгружает пользователю файл шаблона импорта шаблон формируется по модели полученной из запроса.
        Т.к. вьюха общая для всех моделей, то файлы генерируются каждый раз при запросе, а не берутся готовые шаблоны.
        Соответственно, файл шаблона, после выгрузки удаляется
    '''

    name_model = request.POST.get('NameRegisterModel', 'Модель не найдена')
    model_for_file = get_model('AppGtn', name_model)

    path_to_template_import = model_for_file.get_file_template_for_import()

    filename = name_model + str(random.randint(1000, 9999)) + '.xlsx'

    return get_response_to_unload_file(path_to_template_import, name_file_for_user=filename, remove_file=True)

def appgtn_import_in_system (request):

    """
        Вьюха берет данные из файла, опеделяет модель для импорта и импортирует данные в систему
    """

    name_model = request.POST.get('NameRegisterModel', 'Модель не найдена')
    model_for_file = get_model('AppGtn', name_model)

    file_data = request.FILES['file_to_import']

    path_work_file = settings.MEDIA_ROOT+file_data.name
    with open (path_work_file,'wb') as work_file:
        for chunk in file_data.chunks():
            work_file.write(chunk)

    try:
        result_import = model_for_file.import_in_system(path_work_file)
    except RuntimeError as result_import:
        return render(request,'template_test.html',{'parametr':result_import})
    finally:
        os.remove(path_work_file)

    return  render(request,'result_import.html',
                   {'status_import':result_import['message'],
                    'number_of_records':result_import['number_of_records'],
                    'number_of_imported':result_import['number_of_imported'],
                    'number_of_errors':result_import['number_of_errors'],
                    'number_of_duplicates':result_import['number_of_duplicates'],
                    'link_to_file_errors':result_import['link_to_file_errors'],
                    'path_back':DICT_URL_PATH_TO_REGISTER[name_model]
                    })


def appgtn_unload_file_with_import_errors (request):

    """
        Вьюха отадает пользователю файл с ошибками импорта данных в Систему.
        Нужный файл определяется по значению "PathToFileError" из запроса.

        :param request: поступивший запрос
        :return: объект HttpResponse содержащий запрашиваемый файл в двоичном режиме
    """
    return get_response_to_unload_file(request.POST['PathToFileError'])


def appgtn_del_file_and_redirect (request):

    """
        Вюха удаляет файл на сервере (путь определяется по параметру запроса "PathToFileError").
        Далее вьюха переадресовывает пользователя на страницу указанную в параметре "PathBack"

        :param request: поступивший запрос
        :return: Перенаправление на на страницу указанную в параметре "PathBack"
    """

    os.remove(request.POST['PathToFileError'])

    return HttpResponseRedirect(request.POST['PathBack'])


    # #####Попытки переименовать временный файл######
    # # получаем абсолютный путь к файлу
    # abs_file_data = file_data.file.name
    # # f = open (abs_file_data.replace('\\','\\'))
    # # получаем расширение загруженного файла (именно которого грузили, а не которого сохранила Система)
    # file_name,file_extension = file_data.name.rpartition('.')[::2]
    #
    # abs_file_data_path = abs_file_data.rpartition('.')[0]
    #
    # # меняем у загруженного файла расширение с ".upload" на исходный
    # # os.rename(abs_file_data.replace('\\',"/"),str(abs_file_data_path.replace('\\',"/") + '.' + file_extension))
    #
    # os.replace(abs_file_data,abs_file_data_path + '.' + file_extension)
    # #как ни странно, но для получения абсолютного пути к файлу необходимо обратится к атриббуту "file", а затем к "name"
    #
    #
    # result_import = model_for_file.import_in_system(abs_file_data)
    #
    # # result_import = request.FILES['file_to_import'].temporary_file_path
    # # result_import = openpyxl.load_workbook(r'c:\users\sss\appdata\local\temp\krakec.upload'.replace('\\',"/"))
    #
    # ###############################################
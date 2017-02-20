# coding=utf-8
__author__ = 'sss'
'''
    модуль содержит классы и методы выполняющие импорт данных в Систему
    Глобальные переменные объявлены в конце модуля, это сделано для того,
    чтобы в DICT_MATCHING_CLASS_IMPORT_WITH_CLASS_CHECK можно было хранить непосредственно ссылку на класс проверки,
    а не его имя (т.е. чтобы в последующем не проводить поиск по имени класса проверки, полученного из словаря)
'''

import types
import random
import openpyxl
import copy_reg
import multiprocessing

from datetime import datetime
from functools import partial
from django.conf import settings

# _pickle_method и copy_reg необходимы для проведения мультиимпорта
def _pickle_method(m):
    if m.im_self is None:
        return getattr, (m.im_class, m.im_func.func_name)
    else:
        return getattr, (m.im_self, m.im_func.func_name)

copy_reg.pickle(types.MethodType, _pickle_method)



class ImportDataInSystem():

    """
        Класс, который занимается импортом данных и организует проверку данных
    """

    @classmethod
    def get_file_template_for_import(cls, cls_for_template):

        """
            Метод формирует файл-шаблон для дальнейшего заполнения.
            В файл пишуться заголовки столбцов,
                название которых берется из verbose_name

            :param cls_for_template: класс модели для которой осуществляется
                                        формирвоания файла
            :return: полный путь к сформированному файлу
        """

        wb = openpyxl.Workbook()
        ws = wb.active

        j = 1
        for field in cls_for_template._meta.fields:

            if not (field.name in cls_for_template.FIELD_NOT_FOR_IMPORT):
                v_name = field.verbose_name
                param = {'row': 1, 'column': j, 'value': v_name}
                c_letter = openpyxl.utils.get_column_letter(j)

                ws.cell(**param).font = openpyxl.styles.Font(bold=True)
                ws.column_dimensions[c_letter].width = len(v_name) + 1
                j += 1

        # Проводим сохранение файла с данными.
        name_file = ('template_for_import_' + cls_for_template.__name__
                     + str(random.randint(1000, 9999)) + '.xlsx')
        full_name_file = settings.STATIC_ROOT + name_file

        wb.save(full_name_file)

        return full_name_file

    @staticmethod
    def get_true_vlues(cls_for_import, names_check_fields, check_values):

        """
        Метод находит связанные поля в модели, для которой осуществляется импорт.
        Далее метод определяет для найденных связанных полей, по импортируемым данным,
            действительные ("Истинные") значения (в данном случае экземпляр связаной модели целиком)
        Решаемая задача (на примере технического средства (далее - ТС)):
            Импортируется запись о ТС.
            Модель ТС содержит связанное поле "Цвет", которое ссылается на справочник "Цвет ТС".
            В БД, в таблице ТС в колонке "Цвет" хранится ID цвета из справочника "Цвет ТС".
            В тоже время, в импортируемом файле указывается значение цвета, а не его ID (что естетственно, т.к.
                пользователь не знает ID цветов)
            Т.к. типы имортируемых значений для связанных полей и типы связанных полей в БД не совпадают,
                то действовать "в лоб" нельзя.
            Т.о. предварительно необходимо найти по импортируемым значениям соответствующие им экземпляры связанных
                моделей и уже после этого проводить дальнейшую обработку (искать дубли, сохранять данные в БД)

        Нюансы реализации:
            1) Т.к. заренее не известно по какой модели будет проводится импорт, то необходимо проверять все поля модели
               на "связанность"
            2) Аналогично и для связанных моделей:
                т.к. заранее не известна связанная модель, то нельзя определить по какому полю нужно искать
                    импортируемое значение.
                соответственно, необходимо искать импортируемое значение по всем полям модели.
            3) Для корректно сравнения значений проверяемых полей с искомым значением осуществлятся приведение типа:
                приведение типа и сравнение осуществляется на стороне БД, т.к. это менее ресурсоемко чем в Приложении;
            4) Для пустых импортируемых значений (None) поиск не проводится
            5) В результат берем экземпляр связанной модели целиком,
                т.к. этот экземпляр необходимо указывать при сохранении экземпляра импортируемой модели.
            6) Т.к. нет механизма разрешения ситуации когда найденно несколько разных записей с искомым значением,
                то берем первое попавшееся, поэтому в запросе устанавливается ограничеие LIMIT 1

        :param cls_for_import: класс модели для которой осуществляется импорт
        :param names_check_fields: наименование полей модели по которым проводится импорт
        :param check_values: импортируемые значения
        :return: список импортируемых значений, в котором значения для связанных полей заменены соответствующими им
            экземплярами3
        """
        # TODO: https://docs.djangoproject.com/en/1.10/ref/models/meta/ - анализ на упрощение текущего кода
        check_related_generator = (cls_for_import._meta.get_field_by_name(name)[0].rel.to
                                   if cls_for_import._meta.get_field_by_name(name)[0].rel else False
                                   for name in names_check_fields)

        true_check_values = []

        for related_class, value in zip(check_related_generator, check_values):

            true_value = False

            if related_class and value is not None:

                str_column = ','.join(field[0].column for field in related_class._meta.get_fields_with_model())
                str_cast_func = "CAST((%s) as text)" % str_column

                # Ищем связанный объект в связанной модели.
                '''
                    Дефолтный, но не рабочий вариант:
                    related_class.objects.extra(where=['CAST((%s) as text) ~ %s'],params=[str_column,value])
                    Дефолтный вариант не заработал, т.к. в итоговом запросе str_column обернут в дополнительные кавычки,
                       отчего SQL воспринимает str_column как одну строку, а не перечнь наименований колонок.
                       В результате получаем пустой результат.
                       Поэтому функцию приведения типа на стороне БД CAST() оформляем стандартно (str_cast_func)
                '''
                true_value = related_class.objects.extra(where=[str_cast_func +' ~ %s']
                                                         , params=[value])[:1]

            if true_value:
                true_check_values.append(true_value[0])
            else:
                true_check_values.append(value)

        return true_check_values

    @staticmethod
    def check_row_data(class_check, fields_name, check_values):

        """
            Метод проводит проверку данных (check_values) на корректность (согласно логике классов проверки).

            :param class_check: класс проверки, в нем ищуться методы для проверки данных.
            :param fields_name: наименвоание полей, по которым определяется метод проверки.
            :param check_values: проверяемая строка с данными.
            :return: возвращает False если ошибок не найденно, иначе кортеж результатов проверки полей:
            * пустая строка - ошибок в поле нет.
            * текст - поле ошибочно, а полученный текст описывает ошибку.
        """

        result = tuple(getattr(class_check, 'do_check_' + field_name, lambda x: '')(str(check_value))
                                            for field_name, check_value in zip(fields_name, check_values))

        return result if any(result) else False


    @staticmethod
    def check_duplicate(cls_for_import, fields_name, check_values):
        """
            Метод проверки на дублирование записи (check_values) в системе

            :param cls_for_import: класс модели, в которую осуществляется импорт
            :param fields_name: список имен полей, по которым будет проводится поиск дублей
            :param check_values: список значений для полей (fields_name), по которым будет проводится поиск дублей
            :return: True - если дубль, иначе False
        """

        param_for_filter = {key:datetime.date(value) if isinstance(value, datetime) else value
                                        for key, value in zip(fields_name, check_values) if value}

        return True if cls_for_import.objects.filter(**param_for_filter)[:1].exists() else False

    @classmethod
    def check_structure(cls, cls_for_import, row_with_headers):

        """
            Метод проверяет совпадают ли наименования зоголовков из row_with_headers c name или verbose_name полей
             импортируемой модели (cls_for_import).
            Возвращает словарь с двумя списками: в одном перечень полей модели, для которых найденно совпадение среди
            заголовков, в другом - список не идентифицированных загловков.

            :param cls_for_import: класс модели, для которой проводится импорт.
            :param row_with_headers: список с заголовками, которые необходимо проверить.
            :return: возвращается словрь, со следующими ключами:
                * names_check_fields - список name полей модели, для которых найденно совпадение среди заголовков.
                * error_column - список не идентифировацнных заголовков.
        """

        field_for_import = tuple(field[0] for field in cls_for_import._meta.get_fields_with_model()
                                                    if field[0].name not in cls_for_import.FIELD_NOT_FOR_IMPORT)

        names_fields_for_import = tuple(field.name for field in field_for_import)

        verbose_name_to_import = {field.verbose_name: field.name for field in field_for_import}

        error_column = []
        names_check_fields = []
        for cell in row_with_headers:

            if cell.value in names_fields_for_import:
                names_check_fields.append(cell.value)

            elif cell.value in verbose_name_to_import:
                names_check_fields.append(verbose_name_to_import[cell.value])

            else:
                error_column.append(cell.value)

        return {'names_check_fields': names_check_fields, 'error_column': error_column}

    @classmethod
    def do_mono_import(cls, cls_for_import, class_check, names_check_fields, rows_for_import):

        """
            Метод проводит проверку данных (на корректность и дублирование) и пишет данные в БД если нет ошибок.
            Данные для проверки беруться из rows_for_import
            :param cls_for_import: класс, для которого проводим импорт
            :param rows_for_import: импортируемые записи, которые необходимо проверить.
            :return: Список, содержащий результаты проверки записей:
                * Кортеж - в некоторых проверяемых ячейках данных найденны ошибки, описание которых указано
                    в соответствующем элементе кортежа
                * Строка юникода - проверяемая запись является дублем (текст об этом и говорит :) )
                * None - ошибок нет, проверяемая запись записана в базу данных
        """

        result_check_data = []
        for check_row in rows_for_import:

            # Проверям строку на ошибки в данных.
            result_row = None
            result_check = cls.check_row_data(class_check, names_check_fields, check_row)

            if result_check:
                result_row = result_check
            else:
                # Иначе проводим проверку на дублирование.
                true_check_values = cls.get_true_vlues(cls_for_import, names_check_fields, check_row)
                if cls.check_duplicate(cls_for_import, names_check_fields, true_check_values):
                    result_row = u'Дубль записи в Системе'
                else:
                    # Иначе считаем запись корректной и сохраняем ее в БД.
                    instance_for_save = cls_for_import()
                    for name_field, t_value in zip(names_check_fields, true_check_values):
                        if t_value:
                            setattr(instance_for_save, name_field, t_value)

                    instance_for_save.save()

            result_check_data.append(result_row)

        return result_check_data

    @classmethod
    def do_multi_import(cls, cls_for_import, class_check, names_check_fields, rows_for_import, n_cpu):

        """
            Метод организует паралелльное (в несколько процессов) импортирование данных
                и собирает результат
            :param cls_for_import: класс модели, для которой осуществляется импорт
            :param class_check: объект класса, который проводит проверку импортируемых данных
            :param names_check_fields: наименования проверяемых полей импортируемой модели
            :param rows_for_import: данные для импорта
            :param n_cpu: количество ЦПУ (количество процессов = количеству ЦПУ)
            :return: возвращает список, каждый элемент которого это результат проверки
                        соответствующей строки из rows_for_import
        """

        # Разбиваем импортируемые записи на группы(срезы) для передачи отдельным процессам.
        list_data_for_process = []
        row_count = len(rows_for_import)
        div_row, mod_row = divmod(row_count, n_cpu)
        n_from = 0
        n_before = div_row

        for _ in range(n_cpu - 1):
            import_slice = rows_for_import[n_from:n_before]
            list_data_for_process.append(import_slice)
            n_from += div_row
            n_before = n_from + div_row

        # Записываем остатки данных в пулл данных для последнего процесса
        n_before += mod_row
        import_slice = rows_for_import[n_from:n_before]
        list_data_for_process.append(import_slice)

        # Проводим проверку данных в нескольких процессах
        pool = multiprocessing.Pool(n_cpu)
        if __name__ == 'AppGtn.register.functionality.imort_data':

            result_check = pool.map(partial(cls.do_mono_import
                                            , cls_for_import
                                            , class_check
                                            , names_check_fields)
                                    , list_data_for_process)

            pool.close()
            pool.join()

        # Объединяем результаты работы процессов в одномерный список
        aggregate_result_check = []
        for result_check_slice in result_check:
            aggregate_result_check.extend(result_check_slice)

        return aggregate_result_check

    @classmethod
    def multiprocessor_do_import(cls, cls_for_import, data_file):

        """
            Метод из полученного файла загружает данные в систему;
             При этом проводит проверки на соответствие:
             - расширению
             - формату
             - отсутствию дублей
             - корректности значений полей файла (необходимо настраивать для каждого реестра отдельно)

             В результате возвращается cловарь с итогами импорта и путем к файлу с ошибками

            :param cls_for_import: класс модели, для которой осуществляется импорт
            :param data_file: файл с данными для импорта
            :return: cловарь с итогами импорта и путем к файлу с ошибками
        """

        dict_result = {
            'status': '',
            'message': '',
            'number_of_records': 0,
            'number_of_imported': 0,
            'number_of_errors': 0,
            'number_of_duplicates': 0,
            'link_to_file_errors': ''
        }

        wb = openpyxl.load_workbook(data_file, read_only=False)
        ws = wb.active

        # Проверяем файл на соответствие структуре импорта (есть ли нужные колонки).
        dict_check_structure = cls.check_structure(cls_for_import, ws.rows[0])

        error_column = dict_check_structure['error_column']

        if error_column:

            import_status = 'Fatal error'
            error_message = (u'Импорт невозможен. Не удалось идентифициоровать все колонки в импортируемом файле.'
                                 u' Перечень не идентифицирвоанных колонок: ')

            for error in error_column:
                error_message += error + ','

            dict_result.update({
                'status': import_status,
                'message': error_message
                })

            return dict_result

        check_rows = ws.rows[1:]
        count_records = len(check_rows)
        cpu_count = multiprocessing.cpu_count() - 1
        names_check_fields = dict_check_structure['names_check_fields']
        class_check = DICT_MATCHING_CLASS_IMPORT_WITH_CLASS_CHECK[cls_for_import.__name__]

        # Получаем списки проверяемых значений для каждой строки.
        rows_checked_values = [[cell.value for cell in row] for row in check_rows]

        param_func = [cls_for_import, class_check, names_check_fields, rows_checked_values]
        # В зависимости от количества обрабатываемых записей выполняем проверку или в одном процессе или в нескольких.
        # Граничное число 500 определенно эмпирически, но приблизительно :).
        # Нет смысла распределять импорт по процессам, если их меньше 3.
        if count_records < 500 or cpu_count < 3:
            result_check = cls.do_mono_import(*param_func)
        else:
            param_func.append(cpu_count)
            result_check = cls.do_multi_import(*param_func)

        count_error = 0
        count_duplicates = 0
        # Если проверка выявила хотя бы одну ошибку, то пишем данные в файл с ошибками импорта.
        if any(result_check):

            path_file_error = cls.get_file_template_for_import(cls_for_import)
            wb_error = openpyxl.load_workbook(path_file_error, read_only=False)
            ws_error = wb_error.active

            column_for_write = len(names_check_fields) + 1
            ws_error.cell(row=1, column=column_for_write, value='Описание ошибки')

            # Определяем шаблон заливки ошибочных ячеек.
            redFill = openpyxl.styles.PatternFill(start_color='FFFF0000', fill_type='solid')

            # Пишем описание ошибок в файл и отмечаем цветом ошибочные ячейки.
            for idx_row, result_check_row in enumerate(result_check):

                if result_check_row is None:
                    continue
                else:
                    # Определяем текст ошибки.
                    if isinstance(result_check_row, unicode):
                        text_error = result_check_row
                        count_duplicates += 1
                    else:
                        text_error = ''.join(result_check_row)
                        count_error += 1

                    # В файл, в столбец для записи пишем ошибку.
                    ws_error.cell(row=idx_row+2, column=column_for_write, value=text_error)

                    # Переносим из проверяемого файла в файл для ошибок ошибочную строку.
                    for idx_column, check_cell in enumerate(check_rows[idx_row]):
                        cell_error = ws_error.cell(row=idx_row+2, column=idx_column+1, value=check_cell.value)

                        # Если есть описание ошибки для конкретной ячейки то отмечаем ее цветом.
                        if isinstance(result_check_row, tuple) and result_check_row[idx_column]:
                            cell_error.fill = redFill

            wb_error.save(path_file_error)

        # Анализируем о оформляем результат импорта.
        count_correct_records = count_records - count_error - count_duplicates

        if (count_error or count_duplicates) and count_correct_records:
            import_status = 'Error'
            error_message = (u'Данные загружены частично.'
                             u' В импортируемых данных обнаружены записи с ошибками или дубликаты.')

        elif (count_error or count_duplicates) and not count_correct_records:
            import_status = 'Error'
            error_message = (u'Нет данных для загрузки.'
                             u' В импортируемых данных обнаружены записи с ошибками или дубликаты.')
        else:
            import_status = u'Успех.'
            error_message = u'Загружены все записи.'

        dict_result.update({
            'status': import_status,
            'message': error_message,
            'number_of_records': count_records,
            'number_of_imported': count_correct_records,
            'number_of_errors': count_error,
            'number_of_duplicates': count_duplicates,
            'link_to_file_errors': path_file_error
        })

        return dict_result

class CheckImportEntityNaturalPerson ():
    """
        Класс, который отвечает за проверку полей класса о физических лицах
    """

    @staticmethod
    def do_check_serial_dul(check_value):

        """
        Проверяем серию удостоверения личности - две проверки:
        1) проверка количества символов (должно быть 4)
        2) проверка типа символов - все должны быть цифрами
        """
        result_check = ''

        if len(check_value) != 4:
            result_check += 'Количество символов в серии ДУЛ отлично от 4;'

        if not check_value.isdigit():
            result_check += 'Серия ДУЛ содержит символы не являщиеся цифрами;'

        return result_check

    @staticmethod
    def do_check_number_dul(check_value):

        """
        Проверяем номер удостоверения личности - две проверки:
        1) проверка количества символов (должно быть 6)
        2) проверка типа символов - все должны быть цифрами
        """
        result_check = ''

        if len(check_value) != 6:
            result_check += 'Количество символов в номере ДУЛ отлично от 6;'

        if not check_value.isdigit():
            result_check += 'Номер ДУЛ содержит символы не являющиеся цифрами;'

        return result_check

    @staticmethod
    def do_check_date_issue_dul(check_value):
        """
        Проверяем дату выдачи удостоверения личности - две проверки:
        1) проверка количества символов (должно быть 10)
        2) проверка типа символов - должно удовлетворять двум шаблонам:
        ЦЦЦЦ-ЦЦ-ЦЦ
        ЦЦ.ЦЦ.ЦЦЦЦ
        """

        result_check = ''

        try:

            datetime.strptime(check_value, '%Y-%m-%d %H:%M:%S').date()

        except ValueError:
            result_check += ('Формат записи даты выдачи ДУЛ не соответствует одному из следующих шаблонов: ДД.ММ.ГГГГ '
                             'или ГГГГ-ММ-ДД;')

        return result_check

class CheckImportEntityVehicle():

    """
        Класс, который отвечает за проверку полей класса о юридических лицах
    """
    pass


# TODO: добавить """ Класс, который отвечает за проверку полей класса о технических средствах """.

class CheckImportVehicleColor():
    """
        Класс, который отвечает за проверку полей класса о цветах техники
    """
    pass


# TODO: добавить """ Класс, который отвечает за проверку полей класса о марках двигателя """.

# TODO: добавить """ Класс, который отвечает за проверку полей класса о марках технических средств """.

# TODO: добавить """ Класс, который отвечает за проверку полей класса о выдавшей организации """.


# Глобавльный словарь который сопоставляет какой класс, какую модель проверяет.
DICT_MATCHING_CLASS_IMPORT_WITH_CLASS_CHECK = {'EntityNaturalPerson': CheckImportEntityNaturalPerson,
                                               'VehicleColor': CheckImportVehicleColor,
                                               'EntityVehicle': CheckImportEntityVehicle}